"""
Load structured data (accounts, orders, tickets) from ParcelPilot_Assessment_Data.xlsx
into Postgres. Idempotent: safe to re-run.
"""

import os
import sys
from pathlib import Path
import openpyxl
from datetime import datetime
from sqlalchemy import text
from config import engine

# Adjust path to find xlsx
XLSX_PATH = Path(__file__).parent.parent.parent / "ParcelPilot_Assessment_Data.xlsx"

def load_accounts(ws):
    """Load accounts sheet."""
    accounts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        accounts.append({
            "account_id": row[0],
            "account_name": row[1],
            "plan": row[2],
            "status": row[3],
            "csm": row[4],
            "contract_file": row[5],
            "premium_support": row[6] or False,
        })
    return accounts

def load_orders(ws):
    """Load orders sheet."""
    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        orders.append({
            "order_id": row[0],
            "account_id": row[1],
            "carrier": row[2],
            "status": row[3],
            "booked_at": row[4],
            "pickup_window_start": row[5],
            "pickup_window_end": row[6],
            "pickup_actual_at": row[7],
            "shipment_fee_inr": row[8],
            "carrier_fault": row[9] or False,
            "customer_fault": row[10] or False,
            "cancellation_requested_at": row[11],
            "notes": row[12],
        })
    return orders

def load_tickets(ws):
    """Load tickets sheet."""
    tickets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        tickets.append({
            "ticket_id": row[0],
            "account_id": row[1],
            "created_at": row[2],
            "status": row[3],
            "subject": row[4],
            "description": row[5],
            "channel": row[6],
            "assigned_to": row[7],
            "last_customer_message_at": row[8],
            "historical_resolution": row[9],
        })
    return tickets

def insert_data(table: str, rows: list):
    """Insert rows into table, skipping duplicates."""
    if not rows:
        return

    # Build INSERT statement
    columns = list(rows[0].keys())
    placeholders = ", ".join([f":{col}" for col in columns])
    cols_str = ", ".join(columns)

    sql = f"""
    INSERT INTO {table} ({cols_str})
    VALUES ({placeholders})
    ON CONFLICT DO NOTHING
    """

    with engine.connect() as conn:
        for row in rows:
            conn.execute(text(sql), row)
        conn.commit()

    print(f"✓ Loaded {len(rows)} rows into {table}")

def main():
    if not XLSX_PATH.exists():
        print(f"✗ {XLSX_PATH} not found")
        sys.exit(1)

    print(f"Loading from {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH)

    # Load each sheet
    accounts = load_accounts(wb["accounts"])
    insert_data("accounts", accounts)

    orders = load_orders(wb["orders"])
    insert_data("orders", orders)

    tickets = load_tickets(wb["tickets"])
    insert_data("tickets", tickets)

    print("\n✓ All data loaded successfully")

if __name__ == "__main__":
    main()
